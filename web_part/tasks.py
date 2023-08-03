import csv
import time

from datetime import datetime
from openpyxl import Workbook, load_workbook
from djangoProject.celery import app
from api.models import YcToBp
from apiSout.models import SoutToAc, SoutFromAc, Employee, CommissionMember, RM, ResultMapSOUT, BadFactor
from web_part.models import Files

from django.db.models import Count
from django.db.models.fields.reverse_related import ManyToOneRel
from django.db.models.fields.related import OneToOneField


def create_model_dict(Model):
    list1 = list()
    list2 = list()
    [list1.append(i.name) if type(i) not in [ManyToOneRel, OneToOneField] else list2.append(i.name) for i in Model._meta.get_fields()]
    result = {Model._meta.get_field(item).verbose_name: item for item in list1}
    if Model.__name__ in ['SoutToAc']:
        result['Relations'] = {item: item for item in list2}
    return result


# YcToBpDict = {
#     'id': 'id',
#     'Тип операции': 'operationType',
#     'Табельный номер': 'tabNum',
#     'ФИО': 'FIO',
#     'Код обучения': 'learnCode',
#     'Стоимость курса': 'courseCost',
#     'Наименование программы обучения': 'eduName',
#     'Продолжительность обучения': 'eduTime',
#     'Ссылка на курс обучения сотрудника': 'eduUrl',
#     'Статус обучения': 'eduStatus',
#     'Результат пр знаний': 'result',
#     'Номер протокола': 'protocolNum',
#     'Дата протокола': 'protocolDate',
#     'Член комиссии 1': 'memberId1',
#     'Член комиссии 2': 'memberId2',
#     'Член комиссии 3': 'memberId3',
#     'Дата удостоверения ': 'certDate',
#     'Номер удостоверения': 'certNum',
#     'Номер ФГИС ': 'FGISNum',
#     'Статус СДО': 'platformStatus',
#     'Документы протокола': 'protocol',
#     'Удостоверение(файл)': 'cert',
# }

ModelsDict = {
    'YcToBp': create_model_dict(YcToBp),
    'SoutToAc': create_model_dict(SoutToAc),
    'SoutfromAc': create_model_dict(SoutFromAc),
    'Employee': create_model_dict(Employee),
    'RM': create_model_dict(RM),
    'CommissionMember': create_model_dict(CommissionMember),
    'ResultMapSOUT': create_model_dict(ResultMapSOUT),
    'BadFactor': create_model_dict(BadFactor),
}


@app.task
def update_info(filename):
    ext = filename.split('.')[-1]
    full_path = f'mediafiles/import/{filename}'
    if ext == 'csv':
        csv_update(full_path)
    elif ext == 'xlsx' or ext == 'xls':
        xlsx_update(full_path)


def csv_update(filename):
    reader = csv.DictReader(open(filename), delimiter=';')
    for row in reader:
        row = dict(row)
        with open('mediafiles/logs/error.log', 'w') as f:
            f.write(str(row))
        instance = YcToBp.objects.filter(id=row['id'])
        if instance.exists():
            instance = instance.first()
            for key, value in row.items():
                if key != 'id':
                    field_name = ModelsDict['YcToBp'][key]
                    if 'Date' in field_name:
                        value = str(value).split(' ')[0]
                        value = str(datetime.strptime(value, '%Y-%m-%d').date().strftime('%d.%m.%Y'))
                    setattr(instance, field_name, value)
            instance.save()


def xlsx_update(filename):
    book = load_workbook(filename=filename)
    sheet = book.active
    for row in range(2, sheet.max_row + 1):
        instance = YcToBp.objects.filter(id=sheet.cell(row=row, column=1).value)
        if instance.exists():
            instance = instance.first()
            for column in range(2, sheet.max_column + 1):
                value = sheet.cell(row=row, column=column).value
                if value is not None:
                    field_name = ModelsDict['YcToBp'][sheet.cell(row=1, column=column).value]
                    if 'Date' in field_name:
                        value = str(value).split(' ')[0]
                        value = str(datetime.strptime(value, '%Y-%m-%d').date().strftime('%d.%m.%Y'))
                    setattr(instance, field_name, value)
            instance.save()


@app.task
def get_info_csv(filename):
    time.sleep(2)
    instances = YcToBp.objects.all()
    a_dict = ModelsDict['YcToBp']
    # extra_models = a_dict.get('Relations').keys()
    add_dict = a_dict.copy()
    with open(f'mediafiles/export/{filename}', 'w', newline='', encoding='Windows-1251') as f:
        # if extra_models:
        #     add_dict.pop('Relations')
        #     fieldnames = list(add_dict)
        #     for extra_model in extra_models:
        #         a_model = getattr(instances.last(), extra_model).all()
        #         if a_model.exists():
        #             with open('mediafiles/logs/error.log', 'w') as g:
        #                 a_model = a_model.first().__class__
        #                 amount = list(a_model.objects.values('soutToAc').annotate(amount=Count('soutToAc')).order_by())[0]['amount']
        #                 for num in range(1, amount+1):
        #                     modelDict = ModelsDict[a_model.__name__]
        #                     for key, value in modelDict.items():
        #                         if key not in ['soutToAc', 'resultMapSOUT']:
        #                             g.write(f'{key}')
        #                             fieldnames.append(f'{a_model._meta.verbose_name.title()} №{num}. {a_model._meta.get_field(value).verbose_name}')
        # else:
        fieldnames = list(a_dict)
        dictKeys = add_dict.keys()
        writer = csv.DictWriter(f, delimiter=';', fieldnames=fieldnames)
        headers = dict()
        for field in fieldnames:
            headers[field] = field
        writer.writerow(headers)
        for instance in instances:
            line = dict()
            for dictKey in list(dictKeys):
                line[dictKey] = getattr(instance, ModelsDict['YcToBp'][dictKey])
            if a_dict.get('Relations'):
                pass
            writer.writerow(line)
    file = Files.objects.get(name=filename)
    file.fileField = f'export/{filename}'
    file.save()


@app.task
def get_info_xlsx(filename):
    time.sleep(2)
    instances = YcToBp.objects.all()
    YcToBpKeys = ModelsDict['YcToBp'].keys()
    book = Workbook()
    sheet = book.active
    headers = dict()
    for key in YcToBpKeys:
        if key != 'Relations':
            headers[key] = key
        else:
            pass
    row = 1
    column = 1
    for header in headers.keys():
        sheet.cell(row=row, column=column, value=header)
        column += 1
    max_column = column
    row = 2
    for instance in instances:
        for col in range(1, max_column):
            cell = sheet.cell(row=1, column=col).value
            value = getattr(instance, ModelsDict['YcToBp'][cell])
            if cell == 'Протокол' or cell == 'Удостоверение':
                value = value.name
            sheet.cell(row=row, column=col, value=value)
        row += 1
    book.save(f'mediafiles/export/{filename}')
    file = Files.objects.get(name=filename)
    file.fileField = f'export/{filename}'
    file.save()
